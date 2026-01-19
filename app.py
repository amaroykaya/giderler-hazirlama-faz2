import os
import re
import sys
import shutil
import threading
import queue
import subprocess
import unicodedata
from copy import copy
from pathlib import Path
from datetime import datetime, timedelta

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import pandas as pd
import pdfplumber
import requests
from lxml import etree
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter


# =========================
# Excel kolonları
# =========================
COL_B = "B"  # prefix
COL_D = "D"  # fatura tarihi
COL_K = "K"  # fatura no (eşleşme)
COL_M = "M"  # KDV'siz TL tutar
COL_N = "N"  # etiket
COL_T = "T"  # USD tutarı (M/U)
COL_U = "U"  # USD kur (TCMB alış)
COL_V = "V"  # kur (kullanılmıyor, geriye dönük uyumluluk için)

VALID_TAGS = ["anten", "endüstriyel", "genel", "üretim", "savunma"]
FOLDER_UNMATCHED = "eşleşmedi"

# Etiket normalizasyon fonksiyonu
def normalize_tag(s: str) -> str:
    """
    Etiket normalizasyonu: Türkçe karakterleri ASCII'ye çevir, combining marks temizle.
    ÜRETİM → uretim, ENDÜSTİRİYEL → endustriyel
    Combining dot (İ/i̇) sorununu çözer.
    """
    if not s:
        return ""

    # 1) Unicode NFKD normalize (harf + combining işaretlere ayır)
    s = unicodedata.normalize("NFKD", str(s))

    # 2) Combining mark'ları (özellikle noktalı i'nin noktası) sil
    s = "".join(c for c in s if not unicodedata.combining(c))

    # 3) Türkçe özel harfleri ASCII'ye indir
    s = s.replace("ı", "i")
    s = s.replace("İ", "i")
    s = s.replace("ş", "s")
    s = s.replace("Ş", "s")
    s = s.replace("ğ", "g")
    s = s.replace("Ğ", "g")
    s = s.replace("ü", "u")
    s = s.replace("Ü", "u")
    s = s.replace("ö", "o")
    s = s.replace("Ö", "o")
    s = s.replace("ç", "c")
    s = s.replace("Ç", "c")

    # 4) lowercase
    s = s.lower()

    # 5) harf ve rakam dışındaki her şeyi sil
    s = re.sub(r"[^a-z0-9]+", "", s)

    return s


def parse_excel_date(date_value):
    """
    Excel'den gelen tarih değerini parse eder.
    Gün.ay.yıl formatını (örn: 01.11.2025) datetime'a çevirir.
    """
    if date_value is None:
        return None
    
    try:
        date_str = str(date_value).strip()
        if not date_str or date_str.lower() in ["", "nan", "none", "null"]:
            return None
        
        # gün.ay.yıl formatını parse et (örn: 01.11.2025)
        if "." in date_str and len(date_str.split(".")) == 3:
            parts = date_str.split(".")
            if len(parts) == 3:
                day, month, year = parts[0].strip(), parts[1].strip(), parts[2].strip()
                try:
                    return datetime(int(year), int(month), int(day))
                except ValueError:
                    pass
        
        # pandas dayfirst=True ile parse et
        d = pd.to_datetime(date_str, dayfirst=True, errors="coerce")
        if pd.notna(d):
            return d.to_pydatetime()
    except Exception:
        pass
    
    return None


def read_tl_cell(cell):
    """
    Excel hücresinden gerçek TL tutarını güvenli okur.
    - Yüzde formatlı hücreleri TL'ye çevirmez (None döner)
    - String ise virgül/nokta temizler
    - Decimal / float ise aynen alır
    - Bilimsel notation (E+) güvenli parse eder
    """
    if cell.value is None:
        return None

    # Hücre yüzde formatlıysa -> gerçek TL olamaz, bu durumda değeri kullanma
    if cell.number_format and "%" in str(cell.number_format):
        return None

    val = cell.value

    if isinstance(val, str):
        v = val.strip().replace(" ", "")
        v = v.replace(".", "").replace(",", ".")
        try:
            return float(v)
        except:
            return None

    try:
        return float(val)
    except:
        return None

# Etiket mapping: normalize edilmiş key → klasör adı
TAG_MAP = {
    "anten": "anten",
    "endustriyel": "endüstriyel",
    "genel": "genel",
    "uretim": "üretim",
    "savunma": "savunma"
}

# Fatura numarası regex desenleri
INVOICE_PATTERNS = [
    r"FATURA\s*(NO|NUMARASI|NO\.|NUMARA)\s*[:\-]?\s*([A-Z0-9\-\/\. ]{5,})",
    r"FATURA\s*#\s*[:\-]?\s*([A-Z0-9\-\/\. ]{5,})",
    r"(INVOICE|INV)\s*(NO|NUMBER|#)\s*[:\-]?\s*([A-Z0-9\-\/\. ]{5,})",
]

# FX_LINE_PAT kaldırıldı - PDF'den döviz arama artık yok


# =========================
# Helpers
# =========================
def col_letter_to_idx(letter: str) -> int:
    letter = letter.upper()
    idx = 0
    for ch in letter:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


def safe_str(x) -> str:
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
    except Exception:
        pass
    return str(x).strip()


def normalize_invoice(s):
    """
    Fatura numarasını normalize eder.
    Excel'den gelen sayısal değerleri (scientific notation, .0 eklemeleri) düzeltir.
    """
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\u00A0", "").replace("\u200b", "")
    s = s.strip().upper()

    # Eğer Excel sayısal okuyup .0 eklediyse sil
    if s.endswith(".0"):
        s = s[:-2]

    # Scientific notation engelle
    if "E+" in s or "E-" in s:
        try:
            s = format(int(float(s)), "d")
        except:
            pass

    # Sadece harf ve rakam bırak
    s = re.sub(r"[^A-Z0-9]", "", s)
    return s


def norm_invoice_no(s: str) -> str:
    """Deprecated: normalize_invoice kullanın."""
    return normalize_invoice(s)


def ensure_folders(out_dir: Path):
    out_dir.mkdir(parents=True, exist_ok=True)
    for t in VALID_TAGS + [FOLDER_UNMATCHED]:
        (out_dir / t).mkdir(parents=True, exist_ok=True)


def pdf_text_extract(pdf_path: Path, max_pages: int = 5) -> str:
    parts = []
    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            for page in pdf.pages[:max_pages]:
                txt = page.extract_text() or ""
                if txt:
                    parts.append(txt)
    except Exception:
        return ""
    return "\n".join(parts).strip()


def ocr_available() -> bool:
    try:
        import pytesseract  # noqa: F401
        from pdf2image import convert_from_path  # noqa: F401
        return True
    except Exception:
        return False


def pdf_ocr_extract(pdf_path: Path, max_pages: int = 3) -> str:
    """
    OCR fallback. Windows'ta Tesseract + Poppler yoksa boş döner.
    """
    try:
        from pdf2image import convert_from_path
        import pytesseract

        images = convert_from_path(str(pdf_path), first_page=1, last_page=max_pages)
        out = []
        for img in images:
            out.append(pytesseract.image_to_string(img, lang="tur+eng"))
        return "\n".join(out).strip()
    except Exception:
        return ""





def find_invoice_no(text: str, log_callback=None) -> str:
    """
    PDF'den fatura numarasını bulur (normalize edilmiş değil, ham değer).
    normalize_invoice ile normalize edilmeli.
    """
    def log(msg: str):
        if log_callback:
            log_callback(msg)
    
    t = (text or "").upper()

    for pat in INVOICE_PATTERNS:
        m = re.search(pat, t, flags=re.IGNORECASE)
        if m:
            cand = m.group(m.lastindex).strip()
            cand = cand.split("\n")[0].strip()
            cand = re.sub(r"[^A-Z0-9\-\/\. ]", "", cand).strip()
            if cand:
                log(f"[DEBUG] Found candidate in PDF: '{cand}'")
            return cand

    # Son çare: uzun alfanumerik aday (A612025000112251 gibi)
    candidates = re.findall(r"\b[A-Z]{0,3}\d{8,}\b", t)
    if candidates:
        candidates.sort(key=len, reverse=True)
        for cand in candidates:
            log(f"[DEBUG] Found candidate in PDF: '{cand}'")
        return candidates[0]

    return ""


# parse_fx_amount fonksiyonu kaldırıldı - PDF'den döviz arama artık yok


def tcmb_usd_rate_prev_business_day(invoice_date: datetime, log_callback=None):
    """
    Fatura tarihinden 1 gün önce başlar, yoksa geriye gider (max 10 iş günü).
    Hafta sonları (Cumartesi, Pazar) atlanır.
    TCMB XML ForexBuying (USD alış kuru).
    """
    def log(msg):
        if log_callback:
            log_callback(msg)
    
    for back in range(1, 11):
        d = invoice_date - timedelta(days=back)
        
        # Hafta sonu kontrolü: Cumartesi (5) veya Pazar (6) ise atla
        weekday = d.weekday()
        if weekday == 5 or weekday == 6:  # Cumartesi veya Pazar
            date_str_normalized = d.strftime("%Y-%m-%d")
            log(f"[TCMB] denendi: {date_str_normalized} → hafta sonu, atlandı")
            continue
        
        # Tarih formatı: YYYY-MM-DD (normalize, log için)
        date_str_normalized = d.strftime("%Y-%m-%d")
        
        # TCMB URL için format: YYYYMM/DDMMYYYY
        yyyymm = d.strftime("%Y%m")
        ddmmyyyy = d.strftime("%d%m%Y")
        url = f"https://www.tcmb.gov.tr/kurlar/{yyyymm}/{ddmmyyyy}.xml"

        try:
            r = requests.get(url, timeout=10)
            if r.status_code != 200:
                log(f"[TCMB] denendi: {date_str_normalized} → HTTP {r.status_code}, atlandı")
                continue
            
            root = etree.fromstring(r.content)
            nodes = root.xpath(".//Currency[@CurrencyCode='USD']/ForexBuying")
            
            if not nodes or not nodes[0].text:
                log(f"[TCMB] denendi: {date_str_normalized} → kur bulunamadı")
                continue
            
            val = nodes[0].text.strip()
            rate = float(val.replace(",", "."))
            log(f"[TCMB] denendi: {date_str_normalized} → bulundu: {rate:.4f}")
            return rate
            
        except Exception as e:
            log(f"[TCMB] denendi: {date_str_normalized} → hata: {e}")
            continue

    log(f"[TCMB] son 10 iş günü içinde kur bulunamadı (fatura tarihi: {invoice_date.strftime('%Y-%m-%d')})")
    return None


def open_folder(path: Path):
    try:
        if os.name == "nt":
            os.startfile(str(path))  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            subprocess.run(["open", str(path)], check=False)
        else:
            subprocess.run(["xdg-open", str(path)], check=False)
    except Exception:
        pass


def add_company_header(ws, log_q: queue.Queue):
    """
    Excel sheet'inin en üstüne 3 satır ekler ve AntSis kurumsal başlığını ekler.
    - A1:N3 aralığı #6396CB rengiyle doldurulur
    - Logo D1:F3 alanına eklenir (varsa)
    """
    def log(msg: str):
        log_q.put(msg)
    
    log("[BASLIK] AntSis kurumsal başlığı ekleniyor...")
    
    # En üste 3 satır ekle
    ws.insert_rows(1, 3)
    
    # Mavi renk tanımı (#6396CB)
    blue_fill = PatternFill(start_color="6396CB", end_color="6396CB", fill_type="solid")
    
    # Satır yüksekliklerini ayarla (başlık için daha yüksek) - önce bunu yap ki logo ölçeklendirme doğru olsun
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 25
    
    # D, E, F sütunlarının genişliklerini ayarla (logo için yeterli alan)
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # A1:N3 aralığını mavi ile doldur
    for row in range(1, 4):
        for col in range(1, 15):  # A=1, N=14, range(1,15) = A'dan N'ye kadar
            cell = ws.cell(row=row, column=col)
            cell.fill = blue_fill
    
    # Logo dosyasını bul ve ekle (EXE uyumlu yol)
    if getattr(sys, 'frozen', False):
        base_dir = Path(sys._MEIPASS)
    else:
        base_dir = Path(__file__).parent
    
    logo_paths = [
        base_dir / "logo.png",
        base_dir / "logo.jpg",
        base_dir / "logo.jpeg",
        base_dir / "antsis_logo.png",
        base_dir / "antsis_logo.jpg",
    ]
    
    logo_path = None
    for path in logo_paths:
        if path.exists():
            logo_path = path
            break
    
    if logo_path:
        try:
            # D1:F3 hücrelerini merge et
            ws.merge_cells('D1:F3')
            
            # Logo ekle
            img = Image(str(logo_path))
            
            # D, E, F sütunlarının toplam genişliğini piksel cinsinden hesapla
            # Excel'de 1 karakter genişliği ≈ 7 pixel, col_width karakter cinsinden
            col_width_d = ws.column_dimensions['D'].width or 15
            col_width_e = ws.column_dimensions['E'].width or 15
            col_width_f = ws.column_dimensions['F'].width or 15
            total_col_width = col_width_d + col_width_e + col_width_f
            
            # Satır yüksekliklerini piksel cinsinden hesapla
            # Excel'de 1 point = 1.33 pixel
            row_height_1 = ws.row_dimensions[1].height or 25
            row_height_2 = ws.row_dimensions[2].height or 25
            row_height_3 = ws.row_dimensions[3].height or 25
            total_height_points = row_height_1 + row_height_2 + row_height_3
            
            # Piksel cinsinden hedef boyutlar
            # Excel'de genişlik: total_col_width * 7 pixel (yaklaşık)
            # Excel'de yükseklik: total_height_points * 1.33 pixel
            target_width_px = total_col_width * 7
            target_height_px = total_height_points * 1.33
            
            # Logo boyutunu hedef alana göre ölçekle (aspect ratio korunarak)
            img_width, img_height = img.width, img.height
            width_ratio = target_width_px / img_width
            height_ratio = target_height_px / img_height
            scale_ratio = min(width_ratio, height_ratio)  # En küçük oranı kullan
            
            img.width = int(img_width * scale_ratio)
            img.height = int(img_height * scale_ratio)
            
            # Logo'yu D1 hücresine ekle
            ws.add_image(img, 'D1')
            
            log(f"[BASLIK] ✅ Logo eklendi: {logo_path} (D1:F3, {img.width}x{img.height}px)")
        except Exception as e:
            log(f"[BASLIK] UYARI: Logo eklenirken hata: {e}")
    else:
        log("[BASLIK] UYARI: Logo dosyası bulunamadı (logo.png, logo.jpg, antsis_logo.png aranıyor)")
    
    # AutoFilter satırını 4. satıra ayarla (başlık 1-3, header 4)
    from openpyxl.utils import get_column_letter
    ws.auto_filter.ref = f"A4:{get_column_letter(ws.max_column)}4"
    
    log("[BASLIK] ✅ Başlık eklendi (A1:N3 mavi #6396CB)")


def apply_excel_formatting(ws, log_q: queue.Queue):
    """
    Excel sheet'ine renk formatlaması uygular.
    - 4. satır A'dan U'ya kadar gri (#E5E7EB)
    - Veri satırları (5'ten son dolu satıra kadar) belirli kolonlara renkler uygular
    """
    def log(msg: str):
        log_q.put(msg)
    
    log("[FORMAT] Excel formatlaması uygulanıyor...")
    
    # Renk tanımları
    gray_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    light_green_fill = PatternFill(start_color="D9F2D9", end_color="D9F2D9", fill_type="solid")
    light_yellow_fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")
    light_blue_fill = PatternFill(start_color="DCEBFA", end_color="DCEBFA", fill_type="solid")
    light_orange_fill = PatternFill(start_color="FFE0B3", end_color="FFE0B3", fill_type="solid")
    light_pink_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    
    # 1) 4. satır A'dan U'ya kadar gri
    for col in range(1, 22):  # A=1, U=21, range(1,22) = A'dan U'ye kadar
        cell = ws.cell(row=4, column=col)
        cell.fill = gray_fill
    
    log("[FORMAT] 4. satır (header) gri renklendirildi (A4:U4)")
    
    # 2) Son dolu satırı bul (veri satırları 5'ten başlar)
    last_data_row = 4  # Header satır 4
    for row in range(5, ws.max_row + 1):
        # Herhangi bir sütunda değer varsa bu satır veri satırıdır
        has_data = False
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                has_data = True
                break
        if has_data:
            last_data_row = row
    
    log(f"[FORMAT] Son veri satırı: {last_data_row}")
    
    # 3) Veri satırlarını renklendir (5'ten last_data_row'a kadar)
    if last_data_row >= 5:
        # Kolon harflerini sayıya çevir
        col_map = {
            'B': 2, 'C': 3, 'F': 6, 'I': 9, 'J': 10, 'K': 11, 'L': 12, 'M': 13, 'N': 14, 'Q': 17
        }
        
        # Renk atamaları
        color_assignments = {
            # B ve K → açık yeşil
            2: light_green_fill,   # B
            11: light_green_fill,  # K
            # C ve J → açık sarı
            3: light_yellow_fill,   # C
            10: light_yellow_fill,  # J
            # F ve N → açık mavi
            6: light_blue_fill,     # F
            14: light_blue_fill,   # N
            # I, M ve Q → açık turuncu
            9: light_orange_fill,   # I
            13: light_orange_fill,  # M
            17: light_orange_fill, # Q
            # L → açık pembe
            12: light_pink_fill,    # L
        }
        
        # Her veri satırını işle
        for row in range(5, last_data_row + 1):
            for col_num, fill_color in color_assignments.items():
                cell = ws.cell(row=row, column=col_num)
                cell.fill = fill_color
        
        log(f"[FORMAT] Veri satırları renklendirildi (5-{last_data_row}): B,K=yeşil, C,J=sarı, F,N=mavi, I,M,Q=turuncu, L=pembe")
    else:
        log("[FORMAT] UYARI: Veri satırı bulunamadı")
    
    log("[FORMAT] ✅ Formatlama tamamlandı")


def calculate_rates_on_output_excel(excel_path: Path, log_q: queue.Queue) -> int:
    """
    Faz-2'nin oluşturduğu output Excel üzerinde kur ve USD tutarı hesaplar.
    
    Akış:
    1) new_excel aç
    2) wb.active (output Excel'de tek sheet - Sheet1) seç
    3) D sütunundan tarih oku (datetime olarak, parse yok)
    4) TCMB'den bir önceki iş günü USD alış kurunu bul
    5) U sütununa kur yaz
    6) T sütununa = M / U yaz
    7) kaydet
    """
    def log(msg: str):
        log_q.put(msg)
    
    log(f"[KUR-HESAP] Excel açılıyor: {excel_path}")
    
    # Faz-2'nin oluşturduğu Excel'i aç (output Excel, tek sheet - Sheet1)
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    log(f"[KUR-HESAP] Sheet: {ws.title}")
    
    fx_rows = 0
    
    # Tüm satırları dolaş (başlık satırları 1-3, header satır 4, data satırları 5'ten başlar)
    for excel_row in range(5, ws.max_row + 1):
        # D sütunu: Fatura tarihi - string, datetime veya serial number olabilir
        date_cell = ws[f"{COL_D}{excel_row}"]
        raw = date_cell.value
        
        d = None
        
        if isinstance(raw, datetime):
            d = raw.date()
        
        elif isinstance(raw, (int, float)):
            try:
                d = from_excel(raw).date()
            except:
                d = None
        
        elif isinstance(raw, str):
            s = raw.strip().replace("\u00a0", "")
            if s and s != "-":
                try:
                    d = datetime.strptime(s, "%d.%m.%Y").date()
                except:
                    d = None
        
        if not d:
            log(f"[KUR-HESAP] Satır {excel_row}: Tarih okunamadı (D sütunu='{raw}')")
            continue
        
        # TCMB fonksiyonu datetime beklediği için date'i datetime'a çevir
        invoice_date = datetime.combine(d, datetime.min.time())
        
        log(f"[KUR-HESAP] Satır {excel_row}: Tarih={d.strftime('%d.%m.%Y')}")
        
        # M sütunu: KDV'siz TL tutar
        m_cell = ws[f"{COL_M}{excel_row}"]
        m_value = read_tl_cell(m_cell)
        
        if m_value is None or m_value <= 0:
            log(f"[KUR-HESAP] Satır {excel_row}: M tutarı geçersiz (M='{m_cell.value}')")
            continue
        
        log(f"[KUR-HESAP] Satır {excel_row}: M={m_value:.2f} TL")
        
        # TCMB'den bir önceki iş günü USD alış kurunu bul
        usd_rate = tcmb_usd_rate_prev_business_day(
            invoice_date, 
            log_callback=lambda msg: log(f"[KUR-HESAP] Satır {excel_row}: {msg}")
        )
        
        if usd_rate is None:
            log(f"[KUR-HESAP] Satır {excel_row}: TCMB'den kur bulunamadı (tarih: {invoice_date.strftime('%d.%m.%Y')})")
            continue
        
        log(f"[KUR-HESAP] Satır {excel_row}: Kur={usd_rate:.4f}")
        
        # U sütununa kur yaz
        ws[f"{COL_U}{excel_row}"].value = float(usd_rate)
        
        # T sütununa = M / U yaz (USD tutarı)
        usd_amount = m_value / usd_rate
        t_cell = ws[f"{COL_T}{excel_row}"]
        t_cell.value = float(usd_amount)
        # T sütununa 2 ondalık formatı ekle (0.00)
        t_cell.number_format = "0.00"
        
        log(f"[KUR-HESAP] Satır {excel_row}: ✅ U={usd_rate:.4f} kur yazıldı, T={usd_amount:.2f} USD yazıldı (M={m_value:.2f} TL / {usd_rate:.4f})")
        fx_rows += 1
    
    # En son veri satırını tespit et (SADECE A sütununa bakarak)
    last_data_row = 4  # Header satır 4 (başlık 1-3, header 4)
    for row in range(5, ws.max_row + 1):
        a_cell = ws[f"A{row}"]
        if a_cell.value is not None:
            last_data_row = row
    
    # Dip toplam ekle: Son veri satırından sonra 3 boş satır, sonra TOPLAM (tek satır)
    if last_data_row > 4:
        total_row = last_data_row + 3  # Son veri satırı + 3 boş satır (toplam satırı)
        data_start = 5  # Veri 5. satırdan başlar
        
        # I, L, M, Q, T kolonlarında formüllü dip toplam
        ws[f"I{total_row}"].value = f"=SUM(I{data_start}:I{last_data_row})"
        ws[f"L{total_row}"].value = f"=SUM(L{data_start}:L{last_data_row})"
        ws[f"M{total_row}"].value = f"=SUM(M{data_start}:M{last_data_row})"
        ws[f"Q{total_row}"].value = f"=SUM(Q{data_start}:Q{last_data_row})"
        ws[f"T{total_row}"].value = f"=SUM(T{data_start}:T{last_data_row})"
        
        # Sayı formatı ekle
        ws[f"I{total_row}"].number_format = "#,##0.00"
        ws[f"L{total_row}"].number_format = "#,##0.00"
        ws[f"M{total_row}"].number_format = "#,##0.00"
        ws[f"Q{total_row}"].number_format = "#,##0.00"
        ws[f"T{total_row}"].number_format = "#,##0.00"
        
        log(f"[KUR-HESAP] Dip toplam eklendi: Satır {total_row}, I{total_row}, L{total_row}, M{total_row}, Q{total_row}, T{total_row}")
    
    # Kaydet
    log(f"[KUR-HESAP] {fx_rows} satır için kur hesaplandı, Excel kaydediliyor...")
    wb.save(excel_path)
    wb.close()
    log(f"[KUR-HESAP] Excel kaydedildi: {excel_path}")
    
    return fx_rows


# =========================
# Core processing
# =========================
def process_all(excel_path: Path, pdf_paths: list[Path], out_dir: Path, log_q: queue.Queue):
    def log(msg: str):
        log_q.put(msg)

    ensure_folders(out_dir)

    log("[INIT] Excel okunuyor...")
    
    # Excel'i openpyxl ile açarak K sütununu string olarak oku
    wb = load_workbook(excel_path, data_only=True)
    ws = wb["Tümü"]
    
    idx_B = col_letter_to_idx(COL_B)
    idx_D = col_letter_to_idx(COL_D)
    idx_K = col_letter_to_idx(COL_K)
    idx_M = col_letter_to_idx(COL_M)
    idx_N = col_letter_to_idx(COL_N)
    
    # Excel K => Excel satır numarası map (openpyxl ile oku, string olarak)
    invoice_to_excel_row: dict[str, int] = {}
    # Excel satır numarası => B sütunu değeri (openpyxl ile oku, string olarak)
    excel_row_to_b_value: dict[int, str] = {}
    
    for excel_row in range(2, ws.max_row + 1):  # 2'den başla (header satır 1)
        # K sütunu (fatura no) okuma
        k_cell = ws[f"{COL_K}{excel_row}"]
        if k_cell.value is not None:
            inv_raw = str(k_cell.value).strip() if k_cell.value else ""
            inv = normalize_invoice(inv_raw)
            if inv:
                invoice_to_excel_row[inv] = excel_row  # Excel satır numarası (1-based)
        
        # B sütunu (prefix) okuma - openpyxl ile string olarak
        b_cell = ws[f"{COL_B}{excel_row}"]
        b_raw = ""
        if b_cell.value is not None:
            b_raw = str(b_cell.value).strip()
        # Boşlukları temizle
        b_clean = " ".join(b_raw.split()) if b_raw else ""
        excel_row_to_b_value[excel_row] = b_clean
    
    # Diğer sütunlar için pandas kullan (dtype=str ile)
    df = pd.read_excel(excel_path, dtype=str)
    
    # Excel satır numarası => DataFrame index map oluştur
    excel_row_to_df_idx: dict[int, int] = {}
    # pandas header=0 varsayılan (ilk satır header), data satırları 1-based Excel'de 2'den başlar
    for df_idx in range(len(df)):
        excel_row = df_idx + 2  # DataFrame index 0 -> Excel satır 2
        excel_row_to_df_idx[excel_row] = df_idx

    can_ocr = ocr_available()
    log(f"[INIT] OCR {'AKTİF' if can_ocr else 'YOK (metinli PDF varsa çalışır)'}")

    total = len(pdf_paths)
    matched = 0

    for n, pdf_path in enumerate(pdf_paths, start=1):
        pdf_filename = os.path.splitext(pdf_path.name)[0]
        log(f"[DEBUG] Processing PDF: {pdf_filename}")
        log(f"[PDF] ({n}/{total}) {pdf_path.name}")

        text = pdf_text_extract(pdf_path, max_pages=5)
        if len(text) < 30:
            if can_ocr:
                ocr_text = pdf_ocr_extract(pdf_path, max_pages=3)
                if len(ocr_text) > len(text):
                    text = ocr_text
            else:
                # OCR yoksa ve metin de yoksa eşleşmedi
                if len(text) < 10:
                    dest = out_dir / FOLDER_UNMATCHED / pdf_path.name
                    shutil.copy2(pdf_path, dest)
                    log("[MATCH] metin yok + OCR yok -> eşleşmedi")
                    log(f"[DEBUG] UNMATCHED PDF: {pdf_filename}")
                    continue

        # 1) PDF içinden gerçek invoice numarasını oku → normalize et → Excel K sütununda ara
        inv_no_raw = find_invoice_no(text, log_callback=log)
        inv_no = normalize_invoice(inv_no_raw) if inv_no_raw else ""
        
        excel_row = None
        
        if inv_no:
            excel_row = invoice_to_excel_row.get(inv_no)
            if excel_row:
                log(f"[DEBUG] Trying Excel row={excel_row} invoice='{inv_no}' vs PDF")
            else:
                # K'da yok -> başarısız say -> filename'e geç
                inv_no = ""
        
        # 2) Bulamazsan: PDF dosya adından invoice çıkar → normalize et → Excel K sütununda ara
        # 2) Bulamazsan: PDF dosya adından invoice çıkar → normalize et → Excel K sütununda ara
        if excel_row is None:
            filename = os.path.splitext(os.path.basename(str(pdf_path)))[0]  # AVA2025000441018
            filename_norm = normalize_invoice(filename)
            log(f"[DEBUG] Filename match try: {pdf_filename}")
            
            excel_row_candidate = invoice_to_excel_row.get(filename_norm)
            if excel_row_candidate:
                log(f"[DEBUG] Trying Excel row={excel_row_candidate} invoice='{filename_norm}' vs PDF (filename)")
                log(f"[DEBUG] Filename matched invoice='{filename_norm}'")
                excel_row = excel_row_candidate
                inv_no = filename_norm
                log(f"[FIX-FILENAME] PDF adı ile eşleşti: {inv_no_raw} -> {filename_norm} (satır {excel_row})")
        
        if excel_row is None:
            dest = out_dir / FOLDER_UNMATCHED / pdf_path.name
            shutil.copy2(pdf_path, dest)
            log(f"[MATCH] Excel'de yok inv={inv_no} -> eşleşmedi")
            log(f"[DEBUG] UNMATCHED PDF: {pdf_filename}")
            continue
        
        # Seçim anı
        log(f"[DEBUG] SELECTED invoice='{inv_no}' from row={excel_row}")
        
        # DataFrame index kontrolü (sadece log/uyarı amaçlı, eşleşmeyi iptal etmez)
        df_idx = excel_row_to_df_idx.get(excel_row)
        if df_idx is None:
            log(f"[UYARI] Excel satır {excel_row} DataFrame mapping'inde bulunamadı (eşleşme devam ediyor)")

        # N sütunu (etiket) - PDF kopyalanmadan hemen önce doğrudan oku ve klasör hesapla
        n_cell = ws[f"{COL_N}{excel_row}"]
        tag_raw = ""
        if n_cell.value is not None:
            tag_raw = str(n_cell.value)
        
        # Normalize etiket: Türkçe karakter sadeleştirme + boşluk temizleme
        tag_key = normalize_tag(tag_raw)
        
        # Mapping sözlüğü ile klasör adını bul
        folder = TAG_MAP.get(tag_key, FOLDER_UNMATCHED)
        
        log(f"[ETIKET] Excel satır={excel_row} raw='{tag_raw}' normalized='{tag_key}' → folder='{folder}'")

        # B sütunu (prefix) - openpyxl mapping'den (string, trim edilmiş)
        b_value = excel_row_to_b_value.get(excel_row, "")
        
        # PDF isimlendirme: B_sütunu + "-" + normalize_invoice(fatura_no) + ".pdf"
        # inv_no zaten normalize_invoice ile normalize edilmiş
        new_name = f"{b_value}-{inv_no}.pdf"
        
        # PDF hedef klasörünü belirle
        if folder == FOLDER_UNMATCHED:
            # Eşleşmedi klasörü: direkt klasöre kopyala
            dest_dir = out_dir / FOLDER_UNMATCHED
        else:
            # Etiket klasörü: {etiket}_fat alt klasörüne kopyala
            dest_dir = out_dir / folder / f"{folder}_fat"
        
        # Klasörü oluştur (yoksa)
        dest_dir.mkdir(parents=True, exist_ok=True)
        
        dest = dest_dir / new_name
        shutil.copy2(pdf_path, dest)

        matched += 1
        log(f"[NAME] B={b_value} inv={inv_no} -> {new_name}")
        log(f"[MATCH] OK inv={inv_no} Excel satır={excel_row} tag={folder} -> {new_name}")

    # Faz-2: Yeni Excel dosyası oluştur (input Excel'in kopyası)
    log("[FAZ-2] Yeni Excel dosyası oluşturuluyor...")
    
    # Çıktı Excel'de sadece tek sekme olacak
    # Aktif sekmenin dışındaki tüm sekmeleri sil
    sheet_names = wb.sheetnames[:]
    active_sheet_name = ws.title
    
    for sheet_name in sheet_names:
        if sheet_name != active_sheet_name:
            wb.remove(wb[sheet_name])
    
    # Aktif sekmenin adını "Sheet1" yap (default tek sheet)
    ws.title = "Sheet1"
    
    # AntSis kurumsal başlığını ekle (en üste 3 satır)
    add_company_header(ws, log_q)
    
    # Excel formatlamasını uygula
    apply_excel_formatting(ws, log_q)
    
    # A sütunundan ay bilgisini al (artık A5 hücresi, çünkü 3 satır eklendi)
    month_cell = ws["A5"]
    month = ""
    if month_cell.value is not None:
        month = str(month_cell.value).strip().upper()
    else:
        log("[FAZ-2] UYARI: A5 hücresi boş, varsayılan isim kullanılıyor")
        month = "AY"
    
    # Output Excel dosya adı: gider_kalemleri_<AY>_2025_faz2.xlsx
    output_filename = f"gider_kalemleri_{month}_2025_faz2.xlsx"
    out_excel = out_dir / output_filename
    wb.save(out_excel)
    log(f"[FAZ-2] Yeni Excel kaydedildi: {out_excel}")

    # Input Excel'i kapat
    wb.close()

    log("--------------------------------------------------")
    log(f"[DONE] PDF toplam: {total}")
    log(f"[DONE] Eşleşen: {matched}")
    log(f"[DONE] Yeni Excel: {out_excel}")

    # Faz-3: Yeni Excel üzerinde kur ve USD hesaplama
    log("[FAZ-3] Kur ve USD hesaplama başlıyor...")
    fx_rows = calculate_rates_on_output_excel(out_excel, log_q)
    
    log("--------------------------------------------------")
    log(f"[DONE] Döviz yazılan satır: {fx_rows}")
    
    # Master Excel için freeze ve viewport ayarları - EN SON işlem (calculate_rates_on_output_excel tamamlandıktan sonra)
    wb_final = load_workbook(out_excel, data_only=False)
    ws_final = wb_final.active
    ws_final.freeze_panes = "A5"
    ws_final.sheet_view.topLeftCell = "A1"
    ws_final.sheet_view.selection[0].activeCell = "A1"
    ws_final.sheet_view.selection[0].sqref = "A1"
    wb_final.save(out_excel)
    wb_final.close()
    
    # Faz-4: Klasör bazlı Excel üretimi
    log("[FAZ-4] Klasör bazlı Excel dosyaları oluşturuluyor...")
    create_folder_excels_from_master(out_excel, out_dir, VALID_TAGS, month, "2025", log_q)
    
    log("--------------------------------------------------")
    log("[DONE] İşlem bitti.")


def clone_and_filter_workbook(master_path: Path, output_path: Path, target_tag: str, log_q: queue.Queue) -> bool:
    """
    Master Excel'i dosya seviyesinde kopyalar ve belirli bir etiket için filtreler.
    
    Args:
        master_path: Master Excel dosya yolu
        output_path: Çıktı Excel dosya yolu
        target_tag: Hedef etiket (N sütununda aranacak)
        log_q: Log kuyruğu
    
    Returns:
        bool: Başarılı ise True
    """
    def log(msg: str):
        log_q.put(msg)
    
    try:
        # 1) Master Excel'i dosya seviyesinde kopyala (shutil.copy)
        log(f"[KLASOR-EXCEL] Master Excel kopyalanıyor: {target_tag}")
        shutil.copy(master_path, output_path)
        log(f"[KLASOR-EXCEL] Dosya kopyalandı: {output_path}")
        
        # 2) Kopya üzerinden openpyxl ile aç
        wb = load_workbook(output_path, data_only=False)
        ws = wb.active
        
        # 3) Son veri satırını bul (dip toplam dahil)
        last_data_row = 4  # Header satır 4
        for row in range(5, ws.max_row + 1):
            # TOPLAM kelimesi içeren satırları bul (dip toplam)
            has_total = False
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and "TOPLAM" in str(cell.value).upper():
                    has_total = True
                    break
            
            if has_total:
                last_data_row = row
                break
            
            # Normal veri satırı kontrolü
            has_data = False
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    has_data = True
                    break
            if has_data:
                last_data_row = row
        
        log(f"[KLASOR-EXCEL] Son veri satırı: {last_data_row} (etiket: {target_tag})")
        
        # 4) 5. satırdan last_data_row'a kadar filtreleme yap
        # Geriye doğru sil ki satır numaraları kaymasın
        rows_to_delete = []
        for row in range(5, last_data_row + 1):
            should_delete = False
            
            # Kontrol 1: N sütunu (Etiket) boş mu?
            n_cell = ws[f"{COL_N}{row}"]
            n_value = ""
            if n_cell.value is not None:
                n_value = str(n_cell.value).strip()
            
            if not n_value:
                should_delete = True
                log(f"[KLASOR-EXCEL] Satır {row} silinecek: Etiket (N sütunu) boş")
            
            # Kontrol 2: K sütunu (Fatura No) boş mu?
            if not should_delete:
                k_cell = ws[f"{COL_K}{row}"]
                k_value = ""
                if k_cell.value is not None:
                    k_value = str(k_cell.value).strip()
                
                if not k_value:
                    should_delete = True
                    log(f"[KLASOR-EXCEL] Satır {row} silinecek: Fatura No (K sütunu) boş")
            
            # Kontrol 3: Herhangi bir hücresinde "TOPLAM" geçiyor mu?
            if not should_delete:
                has_total = False
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value and isinstance(cell.value, str) and "TOPLAM" in str(cell.value).upper():
                        has_total = True
                        break
                
                if has_total:
                    should_delete = True
                    log(f"[KLASOR-EXCEL] Satır {row} silinecek: 'TOPLAM' içeriyor")
            
            # Kontrol 4: Etiket eşleşmesi (sadece yukarıdaki kontroller geçtiyse)
            if not should_delete:
                cell_tag = n_value.lower()
                # Etiket normalizasyonu (master Excel'deki gibi)
                normalized_tag = normalize_tag(cell_tag)
                
                # Hedef etiketi normalize et
                target_tag_normalized = normalize_tag(target_tag)
                
                # Eşleşmiyorsa silinecek satırlar listesine ekle
                if normalized_tag != target_tag_normalized:
                    should_delete = True
            
            if should_delete:
                rows_to_delete.append(row)
        
        # Satırları geriye doğru sil
        for row in reversed(rows_to_delete):
            ws.delete_rows(row)
        
        log(f"[KLASOR-EXCEL] {len(rows_to_delete)} satır silindi (etiket: {target_tag})")
        
        # 5) G ve H sütunlarını sil (G=7, H=8)
        ws.delete_cols(7, 2)
        log(f"[KLASOR-EXCEL] G ve H sütunları silindi (etiket: {target_tag})")
        
        # 6) N'den sonraki kolonları kaldır (N=14, O'dan başla)
        if ws.max_column > 14:  # N=14
            ws.delete_cols(15, ws.max_column - 14)
            log(f"[KLASOR-EXCEL] N'den sonraki kolonlar kaldırıldı (etiket: {target_tag})")
        
        # 6) Son veri satırını yeniden bul (satır silme sonrası)
        last_data_row = 4  # Header satır 4
        for row in range(5, ws.max_row + 1):
            has_data = False
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    has_data = True
                    break
            if has_data:
                last_data_row = row
        
        log(f"[KLASOR-EXCEL] Filtreleme sonrası son veri satırı: {last_data_row} (etiket: {target_tag})")
        
        # 7) Dip toplam ekle (G, J, K kolonları - G ve H silindikten sonra)
        if last_data_row >= 5:
            total_row = last_data_row + 4  # 3 boş satır + toplam satırı
            
            # B kolonuna "TOPLAM" yaz
            ws[f"B{total_row}"].value = "TOPLAM"
            
            # G, J, K kolonlarına toplam formülleri ekle (G ve H silindikten sonra yeni harfler)
            ws[f"G{total_row}"].value = f"=SUM(G5:G{last_data_row})"
            ws[f"J{total_row}"].value = f"=SUM(J5:J{last_data_row})"
            ws[f"K{total_row}"].value = f"=SUM(K5:K{last_data_row})"
            
            # Toplam satırına sayı formatı ekle
            ws[f"G{total_row}"].number_format = "#,##0.00"
            ws[f"J{total_row}"].number_format = "#,##0.00"
            ws[f"K{total_row}"].number_format = "#,##0.00"
            
            log(f"[KLASOR-EXCEL] Dip toplam eklendi: Satır {total_row}, G{total_row}, J{total_row}, K{total_row} (etiket: {target_tag})")
        
        # 8) AutoFilter'ı A4:K4 olarak ayarla (kolon silme işleminden sonra)
        ws.auto_filter.ref = "A4:K4"
        log(f"[KLASOR-EXCEL] AutoFilter A4:K4 olarak ayarlandı (etiket: {target_tag})")
        
        # 9) İlk 4 satırı dondur ve viewport ayarları
        ws.freeze_panes = "A5"
        ws.sheet_view.topLeftCell = "A1"
        ws.sheet_view.selection[0].activeCell = "A1"
        ws.sheet_view.selection[0].sqref = "A1"
        
        # 10) Kaydet
        wb.save(output_path)
        wb.close()
        
        log(f"[KLASOR-EXCEL] ✅ Alt Excel kaydedildi: {output_path} (etiket: {target_tag})")
        return True
        
    except Exception as e:
        log(f"[KLASOR-EXCEL] HATA ({target_tag}): {e}")
        return False


def create_folder_excels_from_master(master_excel_path: Path, base_output_dir: Path, tags: list[str], month_str: str, year: str, log_q: queue.Queue):
    """
    Master Excel'den her etiket klasörü için ayrı Excel dosyaları oluşturur.
    
    Args:
        master_excel_path: Master Excel dosya yolu
        base_output_dir: Ana çıktı klasörü
        tags: Etiket listesi (klasör adları)
        month_str: Ay bilgisi (dosya adı için)
        year: Yıl bilgisi (dosya adı için)
        log_q: Log kuyruğu
    """
    def log(msg: str):
        log_q.put(msg)
    
    log(f"[KLASOR-EXCEL] Master Excel'den klasör bazlı Excel'ler oluşturuluyor...")
    log(f"[KLASOR-EXCEL] Master: {master_excel_path}")
    
    created_count = 0
    
    for tag in tags:
        # Klasör yolunu oluştur
        folder_path = base_output_dir / tag
        
        # Klasör yoksa oluştur
        folder_path.mkdir(parents=True, exist_ok=True)
        
        # Dosya adı: gider_kalemleri_<Ay>_<Yıl>_<etiket>.xlsx
        output_filename = f"gider_kalemleri_{month_str}_{year}_{tag}.xlsx"
        output_path = folder_path / output_filename
        
        # Alt Excel oluştur
        if clone_and_filter_workbook(master_excel_path, output_path, tag, log_q):
            created_count += 1
        else:
            log(f"[KLASOR-EXCEL] UYARI: {tag} için Excel oluşturulamadı")
    
    log(f"[KLASOR-EXCEL] ✅ {created_count}/{len(tags)} klasör Excel'i oluşturuldu")


# =========================
# Tkinter UI
# =========================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Gider Hazırlama Faz-2")
        self.geometry("900x620")

        self.excel_path: Path | None = None
        self.pdf_paths: list[Path] = []
        self.out_dir: Path | None = None

        self.log_q: queue.Queue = queue.Queue()
        self.worker: threading.Thread | None = None

        self._build_ui()
        self._poll_log_queue()
        self._update_checks()

    def _build_ui(self):
        pad = 10
        root = ttk.Frame(self)
        root.pack(fill="both", expand=True, padx=pad, pady=pad)

        # Excel
        f1 = ttk.LabelFrame(root, text="1) Excel Seç")
        f1.pack(fill="x", pady=(0, pad))
        self.excel_label = ttk.Label(f1, text="Seçilmedi")
        self.excel_label.pack(side="left", fill="x", expand=True, padx=pad, pady=pad)
        self.excel_ok = ttk.Label(f1, text="✗", width=3)
        self.excel_ok.pack(side="right", padx=(0, pad), pady=pad)
        ttk.Button(f1, text="Excel Seç", command=self.pick_excel).pack(side="right", padx=(0, pad), pady=pad)

        # PDFs
        f2 = ttk.LabelFrame(root, text="2) PDF'leri Seç (çoklu)")
        f2.pack(fill="x", pady=(0, pad))
        self.pdf_label = ttk.Label(f2, text="Seçilmedi")
        self.pdf_label.pack(side="left", fill="x", expand=True, padx=pad, pady=pad)
        self.pdf_ok = ttk.Label(f2, text="✗", width=3)
        self.pdf_ok.pack(side="right", padx=(0, pad), pady=pad)
        ttk.Button(f2, text="PDF'leri Seç", command=self.pick_pdfs).pack(side="right", padx=(0, pad), pady=pad)

        # Output
        f3 = ttk.LabelFrame(root, text="3) Çıktı Klasörü Seç")
        f3.pack(fill="x", pady=(0, pad))
        self.out_label = ttk.Label(f3, text="Seçilmedi")
        self.out_label.pack(side="left", fill="x", expand=True, padx=pad, pady=pad)
        self.out_ok = ttk.Label(f3, text="✗", width=3)
        self.out_ok.pack(side="right", padx=(0, pad), pady=pad)
        ttk.Button(f3, text="Çıktı Seç", command=self.pick_output).pack(side="right", padx=(0, pad), pady=pad)

        # Actions
        act = ttk.Frame(root)
        act.pack(fill="x", pady=(0, pad))
        self.convert_btn = ttk.Button(act, text="Dönüştür", command=self.start_convert, state="disabled")
        self.convert_btn.pack(side="left")
        self.progress = ttk.Progressbar(act, mode="indeterminate")
        self.progress.pack(side="left", fill="x", expand=True, padx=pad)

        # Log area
        lf = ttk.LabelFrame(root, text="Log")
        lf.pack(fill="both", expand=True)
        self.log_text = tk.Text(lf, wrap="word")
        self.log_text.pack(side="left", fill="both", expand=True)
        scr = ttk.Scrollbar(lf, command=self.log_text.yview)
        scr.pack(side="right", fill="y")
        self.log_text.configure(yscrollcommand=scr.set)

    def _set_check(self, label_widget, ok_widget, is_ok: bool, text: str):
        label_widget.config(text=text)
        if is_ok:
            ok_widget.config(text="✓", foreground="green")
        else:
            ok_widget.config(text="✗", foreground="red")

    def _update_checks(self):
        self._set_check(
            self.excel_label,
            self.excel_ok,
            bool(self.excel_path),
            str(self.excel_path) if self.excel_path else "Seçilmedi"
        )
        self._set_check(
            self.pdf_label,
            self.pdf_ok,
            bool(self.pdf_paths),
            f"{len(self.pdf_paths)} PDF seçildi" if self.pdf_paths else "Seçilmedi"
        )
        self._set_check(
            self.out_label,
            self.out_ok,
            bool(self.out_dir),
            str(self.out_dir) if self.out_dir else "Seçilmedi"
        )

        ready = bool(self.excel_path) and bool(self.pdf_paths) and bool(self.out_dir)
        busy = self.worker is not None and self.worker.is_alive()
        self.convert_btn.config(state=("normal" if ready and not busy else "disabled"))

    def pick_excel(self):
        fp = filedialog.askopenfilename(
            title="Excel seç",
            filetypes=[("Excel", "*.xlsx *.xlsm *.xls")]
        )
        if fp:
            self.excel_path = Path(fp)
        self._update_checks()

    def pick_pdfs(self):
        fps = filedialog.askopenfilenames(
            title="PDF seç (çoklu)",
            filetypes=[("PDF", "*.pdf *.PDF")]
        )
        if fps:
            self.pdf_paths = [Path(p) for p in fps]
        self._update_checks()

    def pick_output(self):
        dp = filedialog.askdirectory(title="Çıktı klasörü seç")
        if dp:
            self.out_dir = Path(dp)
        self._update_checks()

    def log(self, msg: str):
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")

    def _poll_log_queue(self):
        try:
            while True:
                msg = self.log_q.get_nowait()
                self.log(msg)
        except queue.Empty:
            pass
        self.after(150, self._poll_log_queue)

    def start_convert(self):
        if not (self.excel_path and self.pdf_paths and self.out_dir):
            return
        if self.worker and self.worker.is_alive():
            return

        self.log("=== BAŞLADI ===")
        self.progress.start(10)
        self._update_checks()

        def run():
            try:
                process_all(self.excel_path, self.pdf_paths, self.out_dir, self.log_q)
            except Exception as e:
                self.log_q.put(f"[ERROR] {e}")
            finally:
                self.after(0, self._on_done)

        self.worker = threading.Thread(target=run, daemon=True)
        self.worker.start()

    def _on_done(self):
        self.progress.stop()
        self._update_checks()
        if self.out_dir:
            if messagebox.askyesno("Bitti", "İşlem bitti. Çıktı klasörünü açayım mı?"):
                open_folder(self.out_dir)


if __name__ == "__main__":
    App().mainloop()
